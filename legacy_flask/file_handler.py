"""
File Handler Module - CSV/Excel Import and Export
Handles bulk data upload and report generation
"""

import io
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


def parse_csv_file(file_content):
    """
    Parse CSV file content
    
    Args:
        file_content: File content in bytes
    
    Returns:
        list: List of parsed records
    """
    try:
        text_content = file_content.decode('utf-8')
        lines = text_content.strip().split('\n')
        
        if not lines:
            raise ValueError("Empty CSV file")
        
        # Parse CSV
        reader = csv.DictReader(io.StringIO(text_content))
        records = list(reader)
        
        # Standardize column names
        standardized_records = []
        for record in records:
            std_record = {}
            for key, value in record.items():
                key_lower = key.lower().strip()
                
                # Map to standard columns
                if 'age' in key_lower:
                    std_record['patient_age'] = value
                elif 'gender' in key_lower or 'sex' in key_lower:
                    std_record['patient_gender'] = value
                elif 'drug' in key_lower or 'medicine' in key_lower:
                    std_record['suspected_drug'] = value
                elif 'reaction' in key_lower or 'effect' in key_lower or 'symptom' in key_lower:
                    std_record['reaction'] = value
                elif 'severity' in key_lower:
                    std_record['severity'] = value
                elif 'causality' in key_lower:
                    std_record['causality'] = value
                elif 'serious' in key_lower:
                    std_record['seriousness'] = value
                elif 'outcome' in key_lower:
                    std_record['outcome'] = value
                elif 'follow' in key_lower:
                    std_record['follow_up'] = value
                elif 'reporter' in key_lower:
                    std_record['reporter_name'] = value
                elif 'note' in key_lower or 'comment' in key_lower:
                    std_record['notes'] = value
            
            standardized_records.append(std_record)
        
        return standardized_records
    
    except Exception as e:
        raise ValueError(f"Error parsing CSV file: {str(e)}")


def export_to_excel(reports, signals, stats):
    """
    Export analysis results to Excel workbook with multiple sheets and pivot tables
    
    Args:
        reports: List of ADRReport objects
        signals: List of DetectedSignal objects
        stats: Dictionary of statistics
    
    Returns:
        bytes: Excel file content
    """
    output = io.BytesIO()
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # ===== SHEET 1: Summary =====
    ws_summary = wb.create_sheet("Summary", 0)
    
    ws_summary['A1'] = "PHARMACOVIGILANCE ANALYSIS REPORT"
    ws_summary['A1'].font = title_font
    ws_summary.merge_cells('A1:D1')
    
    ws_summary['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    
    # Summary statistics
    row = 4
    ws_summary[f'A{row}'] = "SUMMARY STATISTICS"
    ws_summary[f'A{row}'].font = Font(bold=True, size=12)
    
    row += 1
    ws_summary[f'A{row}'] = "Total ADR Reports:"
    ws_summary[f'B{row}'] = len(reports)
    
    row += 1
    ws_summary[f'A{row}'] = "Total Signals Detected:"
    ws_summary[f'B{row}'] = stats.get('total', 0)
    
    row += 1
    ws_summary[f'A{row}'] = "Strong Signals (PRR > 100):"
    ws_summary[f'B{row}'] = stats.get('strong', 0)
    
    row += 1
    ws_summary[f'A{row}'] = "Moderate Signals (PRR 10-100):"
    ws_summary[f'B{row}'] = stats.get('moderate', 0)
    
    row += 1
    ws_summary[f'A{row}'] = "Weak Signals (PRR 2-10):"
    ws_summary[f'B{row}'] = stats.get('weak', 0)
    
    row += 1
    ws_summary[f'A{row}'] = "Average PRR:"
    ws_summary[f'B{row}'] = stats.get('avg_prr', 0)
    
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 15
    
    # ===== SHEET 2: All Reports =====
    ws_reports = wb.create_sheet("All Reports", 1)
    
    # Headers
    headers = ['ID', 'Patient Age', 'Gender', 'Drug', 'Reaction', 'Severity', 
               'Causality', 'Seriousness', 'Outcome', 'Follow-up', 'Reporter', 'Date']
    
    for col, header in enumerate(headers, 1):
        cell = ws_reports.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
    
    # Add report data
    for row_idx, report in enumerate(reports, 2):
        ws_reports.cell(row=row_idx, column=1).value = report.id
        ws_reports.cell(row=row_idx, column=2).value = report.patient_age
        ws_reports.cell(row=row_idx, column=3).value = report.patient_gender
        ws_reports.cell(row=row_idx, column=4).value = report.suspected_drug
        ws_reports.cell(row=row_idx, column=5).value = report.reaction
        ws_reports.cell(row=row_idx, column=6).value = report.severity
        ws_reports.cell(row=row_idx, column=7).value = report.causality
        ws_reports.cell(row=row_idx, column=8).value = report.seriousness
        ws_reports.cell(row=row_idx, column=9).value = report.outcome
        ws_reports.cell(row=row_idx, column=10).value = report.follow_up
        ws_reports.cell(row=row_idx, column=11).value = report.reporter_name
        ws_reports.cell(row=row_idx, column=12).value = report.report_date.strftime('%Y-%m-%d') if report.report_date else ''
    
    # Adjust column widths
    for col in range(1, len(headers) + 1):
        ws_reports.column_dimensions[get_column_letter(col)].width = 15
    
    # ===== SHEET 3: Detected Signals =====
    ws_signals = wb.create_sheet("Detected Signals", 2)
    
    signal_headers = ['Drug', 'Side Effect', 'PRR', 'Cases', 'Background', 'Strength', 'Confidence %']
    
    for col, header in enumerate(signal_headers, 1):
        cell = ws_signals.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
    
    # Add signal data sorted by PRR
    sorted_signals = sorted(signals, key=lambda x: x.prr_value, reverse=True)
    for row_idx, signal in enumerate(sorted_signals, 2):
        ws_signals.cell(row=row_idx, column=1).value = signal.drug_name
        ws_signals.cell(row=row_idx, column=2).value = signal.side_effect
        ws_signals.cell(row=row_idx, column=3).value = signal.prr_value
        ws_signals.cell(row=row_idx, column=4).value = signal.case_count
        ws_signals.cell(row=row_idx, column=5).value = signal.background_count
        ws_signals.cell(row=row_idx, column=6).value = signal.signal_strength
        
        conf_cell = ws_signals.cell(row=row_idx, column=7)
        conf_cell.value = signal.confidence_score * 100
        conf_cell.number_format = '0.0"%"'
        
        # Color code by strength
        if signal.signal_strength == 'Strong':
            ws_signals.cell(row=row_idx, column=6).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            ws_signals.cell(row=row_idx, column=6).font = Font(color="FFFFFF", bold=True)
        elif signal.signal_strength == 'Moderate':
            ws_signals.cell(row=row_idx, column=6).fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    
    for col in range(1, len(signal_headers) + 1):
        ws_signals.column_dimensions[get_column_letter(col)].width = 18
    
    # ===== SHEET 4: Pivot Analysis (Drug-Effect Matrix) =====
    ws_pivot = wb.create_sheet("Drug-Effect Matrix", 3)
    
    # Create pivot-like table
    ws_pivot['A1'] = "DRUG-EFFECT SIGNAL MATRIX (PRR Values)"
    ws_pivot['A1'].font = title_font
    ws_pivot.merge_cells('A1:E1')
    
    # Build drug and effect lists
    drugs = sorted(set(s.drug_name for s in signals))
    effects = sorted(set(s.side_effect for s in signals))
    
    # Create matrix headers
    for col_idx, effect in enumerate(effects, 2):
        cell = ws_pivot.cell(row=2, column=col_idx)
        cell.value = effect
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
    
    # Create matrix data
    for row_idx, drug in enumerate(drugs, 3):
        ws_pivot.cell(row=row_idx, column=1).value = drug
        ws_pivot.cell(row=row_idx, column=1).font = Font(bold=True)
        
        for col_idx, effect in enumerate(effects, 2):
            # Find signal for this drug-effect pair
            signal_value = None
            for signal in signals:
                if signal.drug_name == drug and signal.side_effect == effect:
                    signal_value = signal.prr_value
                    break
            
            if signal_value:
                cell = ws_pivot.cell(row=row_idx, column=col_idx)
                cell.value = signal_value
                cell.alignment = center_align
                
                # Color code by value
                if signal_value >= 100:
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
                elif signal_value >= 10:
                    cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                elif signal_value >= 2:
                    cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    
    # Adjust column widths
    ws_pivot.column_dimensions['A'].width = 25
    for col in range(2, len(effects) + 2):
        ws_pivot.column_dimensions[get_column_letter(col)].width = 15
    
    # ===== SHEET 5: Severity Distribution =====
    ws_severity = wb.create_sheet("Severity Analysis", 4)
    
    severity_counts = {}
    for report in reports:
        if report.severity:
            severity_counts[report.severity] = severity_counts.get(report.severity, 0) + 1
    
    ws_severity['A1'] = "SEVERITY DISTRIBUTION"
    ws_severity['A1'].font = title_font
    
    row = 3
    ws_severity[f'A{row}'] = "Severity"
    ws_severity[f'B{row}'] = "Count"
    ws_severity[f'A{row}'].fill = header_fill
    ws_severity[f'B{row}'].fill = header_fill
    ws_severity[f'A{row}'].font = header_font
    ws_severity[f'B{row}'].font = header_font
    
    for idx, (severity, count) in enumerate(severity_counts.items(), row + 1):
        ws_severity[f'A{idx}'] = severity
        ws_severity[f'B{idx}'] = count
    
    ws_severity.column_dimensions['A'].width = 20
    ws_severity.column_dimensions['B'].width = 15
    
    # Save to bytes
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()


def generate_signal_summary(signals):
    """
    Generate summary statistics for signals
    
    Args:
        signals: List of DetectedSignal objects
    
    Returns:
        dict: Summary statistics
    """
    if not signals:
        return {
            'total': 0,
            'strong': 0,
            'moderate': 0,
            'weak': 0,
            'avg_prr': 0,
            'max_prr': 0,
            'top_drugs': [],
            'top_effects': []
        }
    
    strong = [s for s in signals if s.signal_strength == 'Strong']
    moderate = [s for s in signals if s.signal_strength == 'Moderate']
    weak = [s for s in signals if s.signal_strength == 'Weak']
    
    # Top drugs and effects
    drug_counts = {}
    effect_counts = {}
    for signal in signals:
        drug_counts[signal.drug_name] = drug_counts.get(signal.drug_name, 0) + 1
        effect_counts[signal.side_effect] = effect_counts.get(signal.side_effect, 0) + 1
    
    top_drugs = sorted(drug_counts.items(), key=lambda x: x[1], reverse=True)[:5]
    top_effects = sorted(effect_counts.items(), key=lambda x: x[1], reverse=True)[:5]
    
    prr_values = [s.prr_value for s in signals]
    
    return {
        'total': len(signals),
        'strong': len(strong),
        'moderate': len(moderate),
        'weak': len(weak),
        'avg_prr': round(sum(prr_values) / len(prr_values), 2) if prr_values else 0,
        'max_prr': round(max(prr_values), 2) if prr_values else 0,
        'top_drugs': top_drugs,
        'top_effects': top_effects
    }
